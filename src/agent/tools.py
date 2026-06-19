"""Agent tools: sql_query, semantic_search, fulltext_search, excel_export."""

import os
import sqlite3
from pathlib import Path

import chromadb
from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook

load_dotenv()

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")
DB_PATH = os.path.join(DATA_DIR, "pipeline.db")
CHROMA_DIR = os.path.join(DATA_DIR, "chromadb")
EXPORTS_DIR = os.path.join(DATA_DIR, "exports")


def _get_embed_client() -> OpenAI:
    return OpenAI(
        api_key=os.getenv("MISTRAL_API_KEY_TRANSCRIBE"),
        base_url=os.getenv("MISTRAL_BASE_URL"),
    )


def sql_query(query: str) -> list[dict]:
    """Execute a read-only SQL query on the messages database.

    Args:
        query: SQL SELECT query to execute.

    Returns:
        List of rows as dictionaries.
    """
    normalized = query.strip().upper()
    if not normalized.startswith("SELECT"):
        return [{"error": "Only SELECT queries are allowed."}]

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.execute(query)
        rows = [dict(row) for row in cursor.fetchall()]
    except sqlite3.Error as e:
        rows = [{"error": str(e)}]
    finally:
        conn.close()
    return rows


def semantic_search(
    query: str, top_k: int = 10, channel: str | None = None, sender: str | None = None
) -> list[dict]:
    """Search messages by semantic similarity.

    Args:
        query: Natural language search query.
        top_k: Number of results to return.
        channel: Optional filter ('Mail' or 'Whatsapp').
        sender: Optional sender filter.

    Returns:
        List of matching documents with metadata and distance.
    """
    embed_client = _get_embed_client()
    chroma_client = chromadb.PersistentClient(path=CHROMA_DIR)
    collection = chroma_client.get_collection("messages")

    # Embed query
    response = embed_client.embeddings.create(model="mistral-embed", input=[query])
    query_embedding = response.data[0].embedding

    # Build where filter
    where = None
    conditions = []
    if channel:
        conditions.append({"channel": channel})
    if sender:
        conditions.append({"sender": {"$eq": sender}})
    if len(conditions) == 1:
        where = conditions[0]
    elif len(conditions) > 1:
        where = {"$and": conditions}

    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=top_k,
        where=where,
        include=["documents", "metadatas", "distances"],
    )

    output = []
    for i in range(len(results["ids"][0])):
        output.append(
            {
                "id": results["ids"][0][i],
                "document": results["documents"][0][i][:500],  # truncate for readability
                "metadata": results["metadatas"][0][i],
                "distance": results["distances"][0][i],
            }
        )
    return output


def fulltext_search(query: str, top_k: int = 20) -> list[dict]:
    """Search messages using FTS5 full-text search (keyword matching).

    Args:
        query: Search terms (supports AND, OR, NOT, phrase "quotes").
        top_k: Maximum number of results to return.

    Returns:
        List of matching messages with id, channel, timestamp, sender, snippet.
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        cursor = conn.execute(
            """SELECT m.id, m.channel, m.timestamp, m.sender, m.receiver,
                      m.subject, snippet(messages_fts, 0, '>>>', '<<<', '...', 64) AS snippet
               FROM messages_fts
               JOIN messages m ON m.rowid = messages_fts.rowid
               WHERE messages_fts MATCH ?
               ORDER BY rank
               LIMIT ?""",
            (query, top_k),
        )
        rows = [dict(row) for row in cursor.fetchall()]
    except sqlite3.Error as e:
        rows = [{"error": str(e)}]
    finally:
        conn.close()
    return rows


def excel_export(data: list[dict], filename: str, columns: list[str] | None = None) -> str:
    """Export data to an Excel file.

    Args:
        data: List of row dictionaries.
        filename: Output filename (without extension).
        columns: Optional ordered list of columns. If None, uses keys from first row.

    Returns:
        Absolute path to the generated .xlsx file.
    """
    os.makedirs(EXPORTS_DIR, exist_ok=True)

    if not data:
        return "Error: No data to export."

    if not columns:
        columns = list(data[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    filepath = os.path.join(EXPORTS_DIR, f"{filename}.xlsx")
    wb.save(filepath)
    return os.path.abspath(filepath)


# Tool definitions for OpenAI function calling
TOOL_DEFINITIONS = [
    {
        "type": "function",
        "function": {
            "name": "sql_query",
            "description": (
                "Execute a read-only SQL SELECT query on the messages database. "
                "Tables: messages (id, channel, timestamp, sender, receiver, cc, bcc, "
                "subject, text, word_count, size_bytes, conversation_id, chat_name), "
                "attachments (id, message_id, path, extension, mimetype, original_filename, size_bytes, "
                "audio_transcription, image_summary, video_summary, extracted_text). "
                "channel is 'Mail', 'Whatsapp', or 'Document'. "
                "Use this for counting, filtering, sorting, aggregations, chronological lists."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "SQL SELECT query to execute."},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "semantic_search",
            "description": (
                "Search messages by semantic similarity for topic/content-based questions. "
                "Use this when asking about what was discussed, finding messages about specific topics."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Natural language search query."},
                    "top_k": {"type": "integer", "description": "Number of results (default 10)."},
                    "channel": {
                        "type": "string",
                        "enum": ["Mail", "Whatsapp", "Document"],
                        "description": "Optional channel filter.",
                    },
                    "sender": {"type": "string", "description": "Optional sender filter."},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "fulltext_search",
            "description": (
                "Search messages using keyword matching (FTS5). Best for finding exact words, "
                "names, phrases. Supports AND, OR, NOT operators and quoted phrases. "
                "Use this when looking for specific terms, names, or exact text matches."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Search terms. Use AND/OR/NOT for boolean logic, quotes for exact phrases.",
                    },
                    "top_k": {"type": "integer", "description": "Max results (default 20)."},
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_export",
            "description": (
                "Export data to an Excel .xlsx file for download. "
                "Use this when the user asks for a table, list, or document as Excel."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "data": {
                        "type": "array",
                        "items": {"type": "object"},
                        "description": "List of row dictionaries to export.",
                    },
                    "filename": {"type": "string", "description": "Output filename without extension."},
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Ordered column names. Uses keys from first row if omitted.",
                    },
                },
                "required": ["data", "filename"],
            },
        },
    },
]

# Map function names to callables
TOOL_MAP = {
    "sql_query": sql_query,
    "semantic_search": semantic_search,
    "fulltext_search": fulltext_search,
    "excel_export": excel_export,
}
